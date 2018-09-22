import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame as df


def excel_output(self):
    print(("#" * int(self.id)) + " " + self.id)

    w_name = str(self.id) + "_elemento.xlsx"

    w_elem = pd.ExcelWriter(w_name)

    Results = ['fax', 'press_y', 'dry', 'mx', 'mome_y', 'cor_y', 'press_z', 'drz', 'mome_z', 'cor_z']

    for name in Results:
        x_ = self.__dict__[name].tolist()
        df(x_).to_excel(w_elem, str(name))
    w_elem.save()

    data_ = load_workbook(w_name)
    wb3 = data_.create_sheet('datos', 0)
    count_e = 1
    for att, value in self.__dict__.items():
        if type(value) == float or type(value) == int or type(value) == str or type(value) == list or type(
                value) == bool:
            wb3.cell(count_e, 1, value=str(att))
            wb3.cell(count_e, 2, value=str(value))
            count_e += 1

    data_.save(w_name)
    return None