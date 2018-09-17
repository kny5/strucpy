"""
Catalogo de clases
"""

import math
import itertools
import numpy as np
from numpy import matlib
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame as df
#from os import mkdir, path
#import sys
#import datetime


class Main:

    Nodos = set()
    Elementos = set()
    ac_nodo = 0
    Max_val = []
    v_c_n = []
    preliminar = set()
    marcos = set()

    def __init__(self):
        pass

    def elementCreator(self):
        pass

    @classmethod
    def marco_analysis(cls, number):
        pass

    def complete_analysis(cls):
        pass

    def d3_plot(cls):
        pass


class Nodo:
    n_id_gen = itertools.count(1)
    counter = 0

    def __init__(self, position, **kwargs):
        self.n_id = str(next(self.n_id_gen))
        self.x = position[0]
        self.y = position[1]
        self.z = position[2]
        self.position = position
        self.ve_parcial = []
        self.n_vcn = []
        self.n_apoyo = []

        for value in list(kwargs.values()):
            if value:
                try:
                    self.n_vcn.append(value[0])
                    self.n_apoyo.append(value[1])
                except TypeError:
                    self.n_vcn.append(0)
                    self.n_apoyo.append(0)
                Main.ac_nodo += 1
                self.ve_parcial.append(Main.ac_nodo)
                Main.Max_val.append(Main.ac_nodo)
            else:
                self.ve_parcial.append(0)
                self.n_apoyo.append(0)

        Main.Nodos.add(self)

        Main.v_c_n += self.n_vcn

    def tolist(self):
        return [self.position[0], self.position[1], self.position[2]]


class Elemento:
    """Propiedades generales de elemento"""

    SCC = 20
    poisson = 0.50

    countrefs = 0
    e_id_gen = itertools.count(1)

    def __init__(self):
        self.id = str(next(self.e_id_gen))
        self.nu = None  # ángulo plano xz
        self.lm = None  # ángulo plano xy
        self.kv = 0  # módulo de reacción vertical
        self.kh = 0  # módulo de reacción horizontal
        self.wy = 0  # carga uniformemente dist, TON/ML y
        self.wz = 0  # carga uniforme... z
        self.aw = 0  # angulo de carga
        self.marco = 0
        self.apoyos = None
        self.start = None
        self.end = None
        self._nodoStart = None
        self._nodoEnd = None
        self.start_conf = {'dx': True, 'dy': True, 'dz': True, 'mx': True, 'my': True, 'mz': True}
        self.end_conf = {'dx': True, 'dy': True, 'dz': True, 'mx': True, 'my': True, 'mz': True}
        self.ve = None
        self.toDeactivate = None
        #Main.preliminar.add(self)

    #def add(self):
        print('updating')
        Main.Elementos.add(self)
        print('updated')

    def remove(self):
        Main.Elementos.remove(self)

    def set_nodes(self):
        if self.start != self.end:
            try:
                if self.id == 1:
                    self._nodoStart = Nodo(self.start, **self.start_conf)
                    self._nodoEnd = Nodo(self.end, **self.end_conf)
                else:
                    for nodo in Main.Nodos:
                        if nodo.position == self.start:
                            self._nodoStart = nodo
                        elif nodo.position == self.end:
                            self._nodoEnd = nodo

                if self._nodoStart is None:
                    self._nodoStart = Nodo(self.start, **self.start_conf)
                if self._nodoEnd is None:
                    self._nodoEnd = Nodo(self.end, **self.end_conf)
            except AttributeError:
                self._nodoStart = Nodo(self.start, **self.start_conf)
                self._nodoEnd = Nodo(self.end, **self.end_conf)

            self.ve = self._nodoStart.ve_parcial + self._nodoEnd.ve_parcial
            self.l = abs((((self.end[0] - self.start[0]) ** 2) +
                          ((self.end[1] - self.start[1]) ** 2) +
                          ((self.end[2] - self.start[2]) ** 2)) ** 0.5)
            plane_xz = (((self._nodoEnd.x - self._nodoStart.x)**2) +
                        ((self._nodoEnd.z - self._nodoStart.z)**2)) ** 0.5

            if plane_xz != 0:
                _nu_ = math.degrees(math.asin((self._nodoEnd.z - self._nodoStart.z) / plane_xz))
                if self._nodoEnd.x - self._nodoStart.x < 0:
                    self.nu = 180 - _nu_
                else:
                    self.nu = _nu_
            else:
                self.nu = 0
            self.lm = math.degrees(math.asin((self._nodoEnd.y - self._nodoStart.y)/self.l))
            self.apoyos = self._nodoStart.n_apoyo + self._nodoEnd.n_apoyo
        else:
            print("Error No Start or End points")

        if self.start == self.end:

            print('values are the same')
            print(self.start, self.end)
            return False

        print(("¬" * int(self.id)) + " " + self.id)

        return None

    def calculations(self):
        _scc = self.SCC
        _poisson = self.poisson
        long = self.l
        deltax = long / _scc
        self.dx = deltax
        a_nu = self.nu
        a_lm = self.lm
        f_a = (self.kv * self.a1() * deltax ** 4) / (1000 * self.e * self.izz())
        f_b = (self.kh * self.a2() * deltax ** 4) / (1000 * self.e * self.iyy())
        mzz = (self.e * self.izz()) / deltax ** 2
        myy = (self.e * self.iyy()) / deltax ** 2
        vzz = (self.e * self.izz()) / (2 * deltax ** 3)
        vyy = (self.e * self.iyy()) / (2 * deltax ** 3)
        axial = (self.e * self.area()) / self.l
        torsion = ((self.e / (2 * (1 + _poisson))) * self.j()) / self.l
        wy = self.wy
        wz = self.wz
        aw = self.aw
        p_mat = self.p_mat
        m_elast = self.e
        self.vzz = vzz
        self.vyy = vyy
        self.myy = myy
        self.mzz = mzz

        def k__(c):
            k = np.matlib.zeros(shape=((_scc + 1), (_scc + 1)))
            k[0, 0] = k[-1, -1] = 3
            k[1, 1] = k[-2, -2] = 7 + c
            k[1, 0] = k[1, 2] = k[-2, -1] = k[-2, -3] = -4
            k[1, 3] = k[-2, -4] = 1
            for i in range(2, _scc - 1):
                k[i, i] = 6 + c
                k[i, i - 2] = k[i, i + 2] = 1
                k[i, i - 1] = k[i, i + 1] = -4
            return k

        kzz = k__(f_a)
        self.kzz = kzz
        kyy = k__(f_b)
        self.kyy = kyy

        try:
            for i, _bool in enumerate(self.toDeactivate, 0):
                if _bool and i != 0 and i != _scc:
                    kzz[i, i] -= f_a
        except TypeError:
            pass

        def imext__(v, *args):
            n1_ = v[1]
            n2_ = -v[2] + (8 * v[1])
            p1_ = v[-2]
            p2_ = (8 * v[-2]) - v[-3]

            v1 = np.append(np.insert(v, 0, (n2_, n1_)), (p1_, p2_))
            for tupl_ in args:
                if len(tupl_) == 3:
                    v1[tupl_[0]] += tupl_[1] * v1[tupl_[2]]
                else:
                    v1[tupl_[0]] += tupl_[1]
            return v1

        d1zz = imext__(np.dot(kzz.I, -(np.insert(np.zeros(_scc), 0, 3))).A1, (0, -6, 2))
        d2zz = imext__(np.dot(kzz.I, -(np.append(np.zeros(_scc), 3))).A1, (-1, -6, -3))
        te1zz = imext__(np.dot(kzz.I, -(np.insert(np.zeros(_scc), 1, 2 * deltax))).A1,
                        (0, 8 * deltax), (1, 2 * deltax))
        te2zz = imext__(np.dot(kzz.I, -(np.insert(np.zeros(_scc), -1, 2 * deltax))).A1,
                        (-1, 8 * deltax), (-2, 2 * deltax))
        d1yy = imext__(np.dot(kyy.I, -(np.insert(np.zeros(_scc), 0, 3))).A1, (0, -6, 2))
        d2yy = imext__(np.dot(kyy.I, -(np.append(np.zeros(_scc), 3))).A1, (-1, -6, -3))
        te1yy = imext__(np.dot(kyy.I, -(np.insert(np.zeros(_scc), 1, 2 * deltax))).A1,
                        (0, 8 * deltax), (1, 2 * deltax))
        te2yy = imext__(np.dot(kyy.I, -(np.insert(np.zeros(_scc), -1, 2 * deltax))).A1,
                        (-1, 8 * deltax), (-2, 2 * deltax))

        def tmv__(vector, m__, v=False):
            x = np.zeros(_scc + 1)
            y = 2
            if not v:
                sums = ((1, 1), (-2, 0), (1, -1))
            else:
                sums = ((1, 2), (-2, 1), (2, -1), (-1, -2))
            for i in range(_scc + 1):
                for tupl_ in sums:
                    x[i] += tupl_[0] * vector[i + y + tupl_[1]]
            z = np.dot(x, m__)
            return z

        m1zz = tmv__(d1zz, mzz)
        m2zz = tmv__(d2zz, mzz)
        tm1zz = tmv__(te1zz, mzz)
        tm2zz = tmv__(te2zz, mzz)

        m1yy = tmv__(d1yy, myy)
        m2yy = tmv__(d2yy, myy)
        tm1yy = tmv__(te1yy, myy)
        tm2yy = tmv__(te2yy, myy)

        v1zz = tmv__(d1zz, vzz, v=True)
        v2zz = tmv__(d2zz, vzz, v=True)
        tv1zz = tmv__(te1zz, vzz, v=True)
        tv2zz = tmv__(te2zz, vzz, v=True)

        v1yy = tmv__(d1yy, vyy, v=True)
        v2yy = tmv__(d2yy, vyy, v=True)
        tv1yy = tmv__(te1yy, vyy, v=True)
        tv2yy = tmv__(te2yy, vyy, v=True)

        keb = np.matlib.zeros(shape=(12, 12))
        # region one
        keb[0, 0] = axial
        keb[1, 1] = - v1zz[0]
        keb[1, 5] = - tv1zz[0]
        keb[2, 2] = - v1yy[0]
        keb[2, 4] = tv1yy[0]
        keb[3, 3] = torsion
        keb[4, 2] = - m1yy[0]
        keb[4, 4] = tm1yy[0]
        keb[5, 1] = m1zz[0]
        keb[5, 5] = tm1zz[0]
        # region two
        keb[6, 0] = - axial
        keb[7, 1] = v1zz[_scc]
        keb[7, 5] = tv1zz[_scc]
        keb[8, 2] = v1yy[_scc]
        keb[8, 4] = - tv1yy[_scc]
        keb[9, 3] = - torsion
        keb[10, 2] = m1yy[_scc]
        keb[10, 4] = - tm1yy[_scc]
        keb[11, 1] = - m1zz[_scc]
        keb[11, 5] = - tm1zz[_scc]
        # region three
        keb[0, 6] = - axial
        keb[1, 7] = - v2zz[0]
        keb[1, 11] = tv2zz[0]
        keb[2, 8] = - v2yy[0]
        keb[2, 10] = - tv2yy[0]
        keb[3, 9] = - torsion
        keb[4, 8] = - m2yy[0]
        keb[4, 10] = - tm2yy[0]
        keb[5, 7] = m2zz[0]
        keb[5, 11] = - tm2zz[0]
        # region four
        keb[6, 6] = axial
        keb[7, 7] = v2zz[_scc]
        keb[7, 11] = - tv2zz[_scc]
        keb[8, 8] = v2yy[_scc]
        keb[8, 10] = tv2yy[_scc]
        keb[9, 9] = torsion
        keb[10, 8] = m2yy[_scc]
        keb[10, 10] = tm2yy[_scc]
        keb[11, 7] = - m2zz[_scc]
        keb[11, 11] = tm2zz[_scc]

        tr = np.matlib.zeros(shape=(12, 12))
        # rotational matrix
        cosnu = math.cos(math.radians(a_nu))
        sinu = math.sin(math.radians(a_nu))
        coslm = math.cos(math.radians(a_lm))
        sinlm = math.sin(math.radians(a_lm))
        # region one
        tr[0, 0] = cosnu * coslm
        tr[0, 1] = - cosnu * sinlm
        tr[0, 2] = - sinu
        tr[1, 0] = sinlm
        tr[1, 1] = coslm
        tr[2, 0] = sinu * coslm
        tr[2, 1] = - sinu * sinlm
        tr[2, 2] = cosnu
        tr[3, 3] = cosnu * coslm
        tr[3, 4] = - cosnu * sinlm
        tr[3, 5] = - sinu
        tr[4, 3] = sinlm
        tr[4, 4] = coslm
        tr[5, 3] = sinu * coslm
        tr[5, 4] = - sinu * sinlm
        tr[5, 5] = cosnu
        # region two
        tr[6, 6] = cosnu * coslm
        tr[6, 7] = - cosnu * sinlm
        tr[6, 8] = - sinu
        tr[7, 6] = sinlm
        tr[7, 7] = coslm
        tr[8, 6] = sinu * coslm
        tr[8, 7] = - sinu * sinlm
        tr[8, 8] = cosnu
        tr[9, 9] = cosnu * coslm
        tr[9, 10] = - cosnu * sinlm
        tr[9, 11] = - sinu
        tr[10, 9] = sinlm
        tr[10, 10] = coslm
        tr[11, 9] = sinu * coslm
        tr[11, 10] = - sinu * sinlm
        tr[11, 11] = cosnu

        kebg = np.dot(np.dot(tr, keb), tr.T)

        try:
            for i, k in enumerate(self.apoyos, 0):
                kebg[i, i] += k
        except TypeError:
            pass

        self.tr = tr

        self.keb = keb

        self.kebg = kebg

        ###########
        pp = self.area() * (p_mat / 10000)

        self.pp_scc = ((pp * (long / 100)) / _scc) * sinlm

        p_axial = ((- sinlm *
                    pp * (long / 100)) / 2) + \
                  ((- math.sin(math.radians(aw)) * wy *
                    (long / 100)) / 2)

        p_scc_y = ((pp * (long / 100) * coslm) / _scc) * \
                  ((deltax ** 3) / (m_elast * self.izz()))

        p_scc_z = 0
        w_scc_y = ((((wy * long) / 100) *
                    math.cos(math.radians(aw))) / _scc) * \
                  ((deltax ** 3) / (m_elast * self.izz()))

        w_scc_z = ((wz * (long / 100)) / _scc) * ((deltax ** 3) / (m_elast * self.iyy()))

        if self.armadura == False:

            vplocal_y = np.insert(np.append(np.full((_scc - 1,), (p_scc_y + w_scc_y)), 0), 0, 0)
            vplocal_z = np.insert(np.append(np.full((_scc - 1,), (p_scc_z + w_scc_z)), 0), 0, 0)

        else:
            p_elem = pp * (long / 100)
            vplocal_y = np.zeros(_scc + 1)
            vplocal_z = np.zeros(_scc + 1)
            p_vertical = -(coslm * p_elem) / 2

        dlzz = np.dot(kzz.I, -vplocal_z).A1
        dlyy = np.dot(kyy.I, -vplocal_y).A1

        self.dlzz = dlzz
        self.dlyy = dlyy

        def v_maker2(dl, m__):
            vector = np.empty(_scc + 1)
            vector[0] = (dl[1] - (2 * dl[0]) + dl[1]) * m__
            vector[-1] = (dl[-2] - (2 * dl[-1]) + dl[-2]) * m__
            for i in range(1, _scc):
                vector[i] = (dl[i + 1] - (2 * dl[i]) + dl[i - 1]) * m__
            return vector

        mdlyy = v_maker2(dlyy, mzz)
        self.mdlyy = mdlyy

        mdlzz = v_maker2(dlzz, myy)
        self.mdlzz = mdlzz

        def v_maker3(dl, v__, vplocal, i__):
            vector = np.empty(_scc + 1)
            vector[0] = ((dl[2] - (2 * dl[1]) + (2 * dl[1]) - ((8 * dl[1]) - dl[2])) * v__) + \
                        ((vplocal[1] / 2) * (m_elast * i__ / (deltax ** 3)))

            vector[1] = (dl[3] - (2 * dl[2]) + (2 * dl[0]) - dl[1]) * v__

            vector[-2] = (dl[-2] - (2 * dl[-1]) + (2 * dl[-3]) - dl[-4]) * v__

            vector[-1] = ((((8 * dl[-2]) - dl[-3]) - (2 * dl[-2]) + (2 * dl[-2]) - dl[-3]) * v__) - \
                         ((vplocal[1] / 2) * (m_elast * i__ / (deltax ** 3)))

            for i in range(2, _scc - 1):
                vector[i] = (dl[i + 2] - (2 * dl[i + 1]) + (2 * dl[i - 1]) - dl[i - 2]) * v__

            return vector

        vdlyy = v_maker3(dlyy, vzz, vplocal_y, self.izz())
        vdlzz = v_maker3(dlzz, vyy, vplocal_z, self.iyy())

        pcu_local = np.zeros(12)

        if self.armadura == False:
            pcu_local[0] = p_axial
            pcu_local[1] = - vdlyy[0]
            pcu_local[2] = - vdlzz[0]
            pcu_local[3] = 0
            pcu_local[4] = - mdlzz[0]
            pcu_local[5] = mdlyy[0]
            pcu_local[6] = p_axial
            pcu_local[7] = vdlyy[_scc]
            pcu_local[8] = vdlzz[_scc]
            pcu_local[9] = 0
            pcu_local[10] = mdlzz[_scc]
            pcu_local[11] = - mdlyy[_scc]
        else:
            pcu_local[0] = p_axial
            pcu_local[1] = p_vertical
            pcu_local[2] = 0
            pcu_local[3] = 0
            pcu_local[4] = 0
            pcu_local[5] = 0
            pcu_local[6] = p_axial
            pcu_local[7] = p_vertical
            pcu_local[8] = 0
            pcu_local[9] = 0
            pcu_local[10] = 0
            pcu_local[11] = 0

        self.pculocal = pcu_local
        self.pc_ = np.dot(tr, pcu_local).A1

        print((">" * int(self.id)) + " " + self.id)

        return None

    def excel_output(self):

        print(("#" * int(self.id)) + " " + self.id)

        w_name = str(self.id) + "_elemento.xlsx"
        #w_namempleteName = path.join(version, w_name)

        w_elem = pd.ExcelWriter(w_name)

        Results = ['fax', 'press_y', 'dry', 'mx', 'mome_y', 'cor_y', 'press_z', 'drz', 'mome_z', 'cor_z']

        for name in Results:
            x_ = self.__dict__[name].tolist()
            df(x_).to_excel(w_elem, str(name))
        w_elem.save()

        data_ = load_workbook(w_name)
        wb3 = data_.create_sheet('datos', 0)
        count_e = 1
        for att, value in self.__dict__.items():
            if type(value) == float or type(value) == int or type(value) == str or type(value) == list or type(
                    value) == bool:
                wb3.cell(count_e, 1, value=str(att))
                wb3.cell(count_e, 2, value=str(value))
                count_e += 1

        data_.save(w_name)


class Concreto(Elemento):
    """Propiedades específicas del concreto"""
    Elemento.countrefs += 1

    def __init__(self):
        Elemento.__init__(self)
        self.b = 0  # ancho de zapata
        self.h = 0  # altura de zapata
        self.b_prima = 0  # ancho de Contratrabe
        self.e = 221.359  # modulo de elasticidad
        self.p_mat = 2.4  # ton/m3

    def a1(self):
        return self.b

    def a2(self):
        return self.h

    def area(self):
        x = self.b_prima * self.h
        return x

    def izz(self):
        x = (self.b_prima * self.h ** 3) / 12
        return x

    def iyy(self):
        x = (self.h * self.b_prima ** 3) / 12
        return x

    def j(self):
        x = ((self.h / 2) * (self.b_prima / 2) ** 3) * \
            ((16 / 3) - (3.36 * ((self.b_prima / 2) / (self.h / 2)) * \
                         (1 - (self.b_prima / 2) ** 4 / (12 * (self.h / 2) ** 4))))
        return x


class Especial(Elemento):
    Elemento.countrefs += 1

    def __init__(self):
        Elemento.__init__(self)
        self.b = 0  # ancho de la sección
        self.h = 0  # altura de la sección
        self.izz_ = 0  # inercia del eje x
        self.iyy_ = 0  # inercia del eje y
        self.j_ = 0  # momento polar de inercia
        self.e = 0  # modulo de elasticidad
        self.p_mat = 0  # peso propio
        # self.a1 = self.b
        # self.a2 = self.h
        self.area_ = 0

    def a1(self):
        return self.b

    def a2(self):
        return self.h

    def area(self):
        return self.area_

    def izz(self):
        return self.izz_

    def iyy(self):
        return self.iyy_

    def j(self):
        return self.j_


class Acero(Elemento):
    def __init__(self):
        Elemento.__init__(self)
        self.e = 0
        self.p_mat = 7.849
        self.armadura = False


class Or(Acero):
    """Propiedades específicas del tipo OR"""
    Elemento.countrefs += 1

    def __init__(self):
        Acero.__init__(self)
        self.d = 0
        self.bf = 0
        self.tf = 0
        self.tw = 0

    def a1(self):
        return self.bf

    def a2(self):
        return self.d

    def area(self):
        x = (self.d * self.bf) - ((self.d - 2 * self.tf) * (self.bf - 2 * self.tw))
        return x

    def izz(self):
        x = 2 * ((self.bf * self.tf ** 3) / 12) + \
            2 * ((self.tw * (self.d - 2 * self.tf) ** 3) / 12) + \
            2 * ((self.bf * self.tf) * ((self.d - self.tf) / 2) ** 2)
        return x

    def iyy(self):
        x = 2 * ((self.d * self.tw ** 3) / 12) + \
            2 * ((self.tf * (self.bf - 2 * self.tw) ** 3) / 12) + \
            2 * ((self.d * self.tw) * ((self.bf - self.tw) / 2) ** 2)
        return x

    def j(self):
        x = (2 * ((self.bf - self.tw) * (self.d - self.tf)) ** 2) / \
            (((self.bf - self.tw) / self.tf) + ((self.d - self.tf) / self.tw))
        return x


class Ir(Acero):
    """Propiedades específicas del tipo IR"""
    Elemento.countrefs += 1

    def __init__(self):
        Acero.__init__(self)
        self.d = 0
        self.bf = 0
        self.tf = 0
        self.tw = 0

    def a1(self):
        return self.bf

    def a2(self):
        return self.d

    def area(self):
        x = (2 * self.bf * self.tf) + ((self.d - 2 * self.tf) * self.tw)
        return x

    def izz(self):
        x = 2 * ((self.bf * (self.tf ** 3)) / 12) + \
            (self.tw * (self.d - (2 * self.tf)) ** 3 / 12) + \
            2 * ((self.bf * self.tf) * ((self.d - self.tf) / 2) ** 2)
        return x

    def iyy(self):
        x = 2 * ((self.tf * self.bf ** 3) / 12) + \
            (((self.d - 2 * self.tf) * self.tw ** 3) / 12)
        return x

    def j(self):
        x = ((2 * self.bf * self.tf ** 3) + ((self.d - self.tf) * self.tw ** 3)) / 3
        return x


class Oc(Acero):
    """Propiedades específicas del tipo OC"""
    Elemento.countrefs += 1

    def __init__(self):
        Acero.__init__(self)
        self.d = 0
        self.t = 0

    def a1(self):
        return self.d

    def a2(self):
        return self.d

    def area(self):
        x = math.pi * ((self.d / 2) ** 2) - (math.pi * ((self.d - (2 * self.t)) / 2) ** 2)
        return x

    def izz(self):
        x = (0.25 * math.pi * (self.d / 2) ** 4) - ((0.25 * math.pi) * ((self.d - 2 * self.t) / 2) ** 4)
        return x

    def iyy(self):
        x = self.izz()
        return x

    def j(self):
        x = ((math.pi * self.d ** 4) / 32) - ((math.pi * (self.d - 2 * self.t) ** 4) / 32)
        return x
