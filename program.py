""" archivo para pruebas """
from pyfiglet import Figlet

f = Figlet(font='slant')
print(f.renderText('Strucpy') + '[MDF]\t[v0.1]\t[2018]' + '\n')

from Model.k_p import calculations as calc
from Model.kest import vdgen
import pandas as pd
from openpyxl import load_workbook
import os
from shutil import copyfile
import sys
from pandas import DataFrame as df
from Model.Classes import Elementos, Nodos

if not os.path.exists("proyectos"):
    os.mkdir("proyectos")

proyecto = sys.argv[1]

path_pro = "proyectos/" + str(proyecto) + "/"

if not os.path.exists(path_pro):
    os.mkdir(path_pro)

version = path_pro + sys.argv[2]

os.mkdir(version)

calc(Elementos)

print('[Elementos]:\t' + '[' + str(len(Elementos)) + ']')
print('[Nodos]:\t' + '[' + str(len(Nodos)) + ']')

vdgen(Elementos)

#map(vdgen, Elementos)

print("\n[Creando archivos de excel]\n")

for k, elemento in enumerate(Elementos, 1):

    print(("#"*k) + " " + str(k))

    w_name = str(k) + "_elemento.xlsx"
    completeName = os.path.join(version, w_name)

    w_elem = pd.ExcelWriter(completeName)

    Results = ['fax', 'press_y', 'dry', 'mx', 'mome_y', 'cor_y', 'press_z', 'drz', 'mome_z', 'cor_z']
    for name in Results:
        x_ = elemento.__dict__[name].tolist()
        df(x_).to_excel(w_elem, str(name))
    w_elem.save()

    data_ = load_workbook(completeName)
    wb3 = data_.create_sheet('datos', 0)
    count_e = 1
    for att, value in elemento.__dict__.items():
        if type(value) == float or type(value) == int or type(value) == str or type(value) == list or type(value) == bool:

          wb3.cell(count_e, 1, value=str(att))
          wb3.cell(count_e, 2, value=str(value))
          count_e += 1

    data_.save(completeName)

copyfile("elementos.py", os.path.join(version, "elementos"))
print("\n[Programa finalizado]\n")
